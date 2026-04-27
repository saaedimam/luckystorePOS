import openpyxl

def fill_item_codes():
    file_path = 'data/inventory/KarbarApp Items Sample.xlsx'
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Headers: ['Item Name', 'Category', 'Sales Price', 'Purchase Price', 'Opening Stock', 'Low Stock', 'Item Code', 'Description']
    # Item Code is column G (index 7 in 1-based, index 6 in 0-based)
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=1):
        # Generate SKU-0001, SKU-0002, etc.
        item_code = f"SKU-{str(row_idx).zfill(4)}"
        ws.cell(row=row_idx + 1, column=7).value = item_code
        
    wb.save(file_path)
    print(f"Filled item codes for {ws.max_row - 1} rows.")

if __name__ == '__main__':
    fill_item_codes()
