import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

def create_excel_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Template"

    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True

    # Define headers
    headers = [
        "name", "barcode", "sku", "category", "supplier", "batch_code", 
        "expiry_date", "cost", "price", "stock_qty", "store_code", "image_url"
    ]

    # Sample data
    sample_rows = [
        [
            "Parachute Coconut Oil 200ml", "1234567890123", "SKU-PO200", "Cosmetics", 
            "Marico Ltd", "BATCH-001", "2026-12-31", 90.00, 120.00, 50, "BR1", 
            "https://example.com/parachute-oil.jpg"
        ],
        [
            "Danish Butter Cookies 300g", "", "SKU-DC300", "Snacks", 
            "Danish Foods", "BATCH-002", "2026-10-15", 180.00, 240.00, 20, "BR1", 
            ""
        ],
        [
            "Sunsilk Pink Shampoo 180ml", "8901234509876", "SKU-SPINK", "Personal Care", 
            "Unilever", "BATCH-003", "2027-01-20", 130.00, 180.00, 35, "BR2", 
            ""
        ]
    ]

    # Style definitions
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=10)
    font_body_bold = Font(name="Segoe UI", size=10, bold=True)
    
    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Steel Blue
    fill_sample = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Very light gray for sample
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Write sample rows
    for row_idx, row_data in enumerate(sample_rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = thin_border
            
            # Alignments & Number formats
            header_name = headers[col_idx-1]
            if header_name in ["cost", "price"]:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif header_name == "stock_qty":
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif header_name in ["barcode", "sku", "store_code", "batch_code", "expiry_date"]:
                cell.alignment = Alignment(horizontal="center")
                # Force barcode & sku as text to prevent scientific notation in Excel
                if header_name in ["barcode", "sku"] and val != "":
                    cell.number_format = "@"

    # Set column widths
    column_widths = {
        "A": 30, # name
        "B": 18, # barcode
        "C": 15, # sku
        "D": 16, # category
        "E": 20, # supplier
        "F": 15, # batch_code
        "G": 14, # expiry_date
        "H": 12, # cost
        "I": 12, # price
        "J": 12, # stock_qty
        "K": 12, # store_code
        "L": 35, # image_url
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze panes below headers
    ws.freeze_panes = "A2"

    # Add data validation drop-downs for Store Code (BR1, BR2, BR3)
    dv_store = DataValidation(type="list", formula1='"BR1,BR2,BR3"', allow_blank=True)
    dv_store.error ='Your entry is not in the list'
    dv_store.errorTitle = 'Invalid Store Code'
    dv_store.prompt = 'Please select a Store Code (e.g. BR1, BR2)'
    dv_store.promptTitle = 'Store Code Selection'
    ws.add_data_validation(dv_store)
    dv_store.add("K2:K1000") # Apply to store_code column

    # Add data validation drop-downs for Category
    dv_cat = DataValidation(
        type="list", 
        formula1='"Grocery,Cosmetics,Snacks,Personal Care,Beverages,Dairy,Bakery,Vegetables,Meat"', 
        allow_blank=True
    )
    dv_cat.error ='Your entry is not in the list'
    dv_cat.errorTitle = 'Invalid Category'
    dv_cat.prompt = 'Please select or type a Product Category'
    dv_cat.promptTitle = 'Category Selection'
    ws.add_data_validation(dv_cat)
    dv_cat.add("D2:D1000") # Apply to category column

    # Ensure output paths exist
    os.makedirs("templates", exist_ok=True)
    os.makedirs("data/samples", exist_ok=True)

    # Save outputs
    path1 = "templates/lucky-store-import-template.xlsx"
    path2 = "data/samples/lucky-store-import-template.xlsx"
    
    wb.save(path1)
    wb.save(path2)
    print(f"Excel templates created successfully at:\n- {path1}\n- {path2}")

if __name__ == "__main__":
    create_excel_template()
